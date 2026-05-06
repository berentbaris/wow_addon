"""
HCE Lore Guide — generates HCE_Lore_Guide.xlsx
Run: pip install openpyxl && python build_lore_sheet.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Enhanced Classes Lore"

HEADER_FILL = PatternFill("solid", fgColor="1a1a2e")
HEADER_FONT = Font(name="Arial", bold=True, color="e6c73f", size=11)
BODY = Font(name="Arial", size=10, color="222222")
BOLD = Font(name="Arial", size=10, color="222222", bold=True)
LINK = Font(name="Arial", size=10, color="4488cc", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")
WRAPC = Alignment(wrap_text=True, vertical="top", horizontal="center")
BDR = Border(
    left=Side("thin", color="CCCCCC"), right=Side("thin", color="CCCCCC"),
    top=Side("thin", color="CCCCCC"), bottom=Side("thin", color="CCCCCC"),
)

FILLS = {
    "Warrior": PatternFill("solid", fgColor="FFF2E6"),
    "Rogue":   PatternFill("solid", fgColor="F2F2E6"),
    "Warlock": PatternFill("solid", fgColor="F0E6F5"),
    "Druid":   PatternFill("solid", fgColor="E6F5E6"),
    "Hunter":  PatternFill("solid", fgColor="E6F0F5"),
    "Shaman":  PatternFill("solid", fgColor="E6F5F0"),
    "Paladin": PatternFill("solid", fgColor="FFF5E6"),
    "Priest":  PatternFill("solid", fgColor="F5F0E6"),
    "Mage":    PatternFill("solid", fgColor="E6E6F5"),
}

headers = ["Enhanced Class", "Base Class", "Spec", "Race", "Gender",
           "Lore Summary", "Key Requirements", "Lore Justification", "Wiki"]
widths  = [22, 12, 16, 12, 10, 48, 42, 58, 38]

for ci, (h, w) in enumerate(zip(headers, widths), 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, WRAPC, BDR
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:I28"
ws.row_dimensions[1].height = 30

DATA = [
# (name, base, spec, race, gender, lore, [(req, justification)...], wiki_url)
("Mountain King", "Warrior", "Full protection tank", "Dwarf", "Male",
 "Mountain Kings represent the martial prowess. They are unconcerned with their race's preoccupation with mechanical devices and mining. They just love a good ale.",
 [("1h mace or axe", "Traditional dwarven weapons — hammers and axes forged in mountain forges"),
  ("Shield", "Mountain Kings are famous for their defensive stances and unbreakable shield walls"),
  ("No professions", "The wiki says they are 'unconcerned with their race's preoccupation with mechanical devices and mining'"),
  ("Flask trinket", "Known for their legendary love of ale — dwarven drinking culture at its peak"),
  ("Drinking beer", "Famous for drinking and using Find Treasure racial ability")],
 "https://warcraft.wiki.gg/wiki/Mountain_king"),

("Brewmaster", "Warrior", "2h arms", "Tauren", "Male",
 "Brewmasters combine martial prowess with brewing mastery. They use staves to fight, and wear iconic Pandaren outfits and rare armor.",
 [("Alchemy", "Core identity — they brew potions and prepare special food concoctions"),
  ("Staff", "Traditional weapon of the wandering brewmaster, used for both combat and stirring brews"),
  ("Exotic gear (can't equip uncommon/green gear)", "Brewmasters wear unconventional, exotic gear suited to their nomadic alchemical lifestyle"),
  ("Lunar Festival suit", "Chen Stormstout's iconic outfit — Brewmasters dress distinctively"),
  ("Drinking during fights", "The Faire's Special Reserve is the most affordable way to stay permanently drunk")],
 "https://warcraft.wiki.gg/wiki/Brewmaster"),

("Demon Hunter", "Warrior", "Dual-wield fury", "Night Elf", "Male",
 "Demon Hunters are night elf renegades who dedicate themselves to destroying demons across Azeroth. They wield twin blades shirtless and operate outside traditional society.",
 [("Dual swords", "Iconic twin-blade combat style of demon hunters like Illidan"),
  ("No chest piece", "Demon hunters fight bare-chested, tattoos serving as their armor"),
  ("Kilt", "Distinctive demon hunter leg armor replacing conventional plate"),
  ("Renegade (can't equip quest rewards)", "Outcasts operating outside normal society — they take nothing from others"),
  ("Anti-demon quests", "Their entire existence is devoted to slaying the Burning Legion")],
 "https://warcraft.wiki.gg/wiki/Demon_hunter"),

("Berserker", "Rogue", "Backstab assassination", "Troll", "Any",
 "Berserkers are Troll warriors who enter states of uncontrollable rage in combat. Steeped in voodoo traditions and alchemy, they use elixirs, potions, and drinks like Thistle Tea to fuel their savage fury.",
 [("Dagger + Sword", "Fast, dual-wielding weapons for rapid berserker strikes"),
  ("Thrown weapons", "Trolls are known for thrown voodoo projectiles and axes"),
  ("Alchemy", "Berserkers brew voodoo potions and Thistle Tea to fuel their rage"),
  ("Grunt (can't equip rare/epic items)", "Berserkers are savage tribals, not refined gear collectors")],
 "https://warcraft.wiki.gg/wiki/Berserker"),

("Warden", "Rogue", "Ghost combat", "Night Elf", "Female",
 "Wardens are elite female night elf jailors and assassins. They operate exclusively in Night Elven zones in Kalimdor. Their owl scouts and thrown glaives are their signature tools.",
 [("Swords", "Wardens wield the Umbra Crescent — a crescent-blade glaive (sword model)"),
  ("Robe", "Wardens wear distinctive hooded cloaks over lighter armor"),
  ("Thrown weapons", "Fan of Knives — the Warden's signature thrown attack"),
  ("Owl companion", "Owl scouts are bonded to Wardens for reconnaissance"),
  ("Homebound", "Oath-sworn to patrol Kalimdor and guard their prisoners")],
 "https://warcraft.wiki.gg/wiki/Warden_(Warcraft_III)"),

("Runemaster", "Rogue", "Hemorrhage subtlety", "Dwarf", "Male",
 "Runemasters are dwarven mystics who inscribe runes across their bodies and gear, and punch their enemies with fist weapons.",
 [("Fist weapons", "Runes are channeled through hand-to-hand magical implements"),
  ("No chest piece", "Runemasters display body-inscribed runes — armor would cover them"),
  ("Kilt", "Traditional dwarven garb for these ancient runic practitioners"),
  ("Enchanting", "Runemasters inscribe runes onto their own gear — a form of enchanting"),
  ("Off-the-shelf gear (only vendor gear)", "They buy plain vendor gear as a blank canvas for runic inscriptions"),
  ("Scrolls", "Scroll Scribes buy and use scrolls as consumable runic buffs")],
 "https://warcraft.wiki.gg/wiki/Runemaster"),

("Pyremaster", "Warlock", "Firestone destruction", "Orc", "Any",
 "Pyremasters are orc ritual-masters who command fire magic. They are associated with orcish shamanic traditions, hence, they ride wolves and don't summon shadow demons or use wands.",
 [("Imp only", "Imps are fire demons — the only appropriate servant for a fire master"),
  ("Melee weaving dagger", "Melee weaving between instant spells requires a quick weapon"),
  ("Firestone", "Fire-attuned spellstones that enhance destruction magic"),
  ("Cooking", "Orcs gather around campfires for ritual feasts and fire ceremonies"),
  ("No wands", "Pyremasters channel fire directly, not through fragile wands"),
  ("Wolf mount", "Traditional orc mount — wolves are their primal allies"),
  ("Campfire to cremate fallen allies", "Light funeral pyres under fallen allies as cremation rituals")],
 "https://warcraft.wiki.gg/wiki/Warlock"),

("Death Knight", "Warlock", "Melee weaving affliction", "Undead", "Male",
 "Death Knights are undead warriors inspired by the original WC2 Death Knights — warlocks souls placed in human knight bodies. They command necrotic magic (drain spells) * skeletal steeds, and hit with their truncheons.",
 [("No demon pet", "Death Knights command undead power, not demons"),
  ("Sword then Pole at 44", "Heavy melee weapons of death — runeblades and halberds"),
  ("No robes / No wands", "Death Knights wear heavy armor, not cloth"),
  ("Skeletal horse mount", "Death Knights ride skeleton steeds bound by necromancy"),
  ("120 AP at 50", "Death Knights are melee powerhouses requiring raw attack power"),
  ("Melee weaving gameplay", "Weave auto-attacks while channeling Drain Life")],
 "https://warcraft.wiki.gg/wiki/Death_knight"),

("Shadowmage", "Warlock", "Shadowbolt demonology", "Gnome", "Female",
 "Shadowmages are warlocks who are outcasts within Alliance society. They have to craft their own gear, and can't make use of the bank. Innkeepers don't give them hearthstones. They always have black cats.",
 [("Tailoring", "Shadowmages craft their own shadow-infused robes"),
  ("Self-made gear", "All equipment must be personally crafted"),
  ("Robe", "Mystical robes woven with shadow essence are mandatory"),
  ("Spellstone", "Shadow-attuned spellstones for enhanced dark casting"),
  ("Black cat companion", "Shadow familiars manifest as mystical black cats"),
  ("Drifter (no hearth/bank)", "Nomadic mages drifting between shadow realms")],
 "https://warcraft.wiki.gg/wiki/Warlock"),

("Druid of the Claw", "Druid", "Feral tank", "Night Elf", "Male",
 "Druids of the Claw are night elf warriors who shapeshift into savage bears, serving as frontline defenders of nature. They rarely go into towns or cities and therefore don't use the amenities. They fight against the Venture Co.",
 [("Ephemeral (can't repair)", "They exist between physical and spiritual realms — gear is transient"),
  ("Drifter (no hearth/bank)", "Wild druids wander the forests without fixed settlements"),
  ("Armored off-hand/weapon/rings", "Nature-strengthened armor that adapts to their feral forms"),
  ("Pro-nature quests", "Defend forests against the Venture Company's exploitation")],
 "https://warcraft.wiki.gg/wiki/Druid_of_the_Claw"),

("Plagueshifter", "Druid", "Deep restoration", "Tauren", "Female",
 "Plagueshifters are Tauren druids who walk the line between healing and decay, using nature magic to fight undead plagues. They seek out Scourge-infested lands to cleanse corruption, carrying jungle remedies as their weapons against blight.",
 [("Partisan (can't equip looted gear)", "They rely on nature's gifts, not spoils of war"),
  ("Jungle Remedy", "Tauren herbalism focuses on tropical plant-based plague cures"),
  ("Anti-scourge quests", "Purifying Scourge corruption across Lordaeron")],
 "https://warcraft.wiki.gg/wiki/Druid"),

("Savagekin", "Druid", "Moonkin balance", "Tauren", "Male",
 "Savagekin are male Tauren druids who embrace primal savagery, channeling both moon and sun magic. They are wild, untamed forces of nature who never leave Kalimdor, bound to the land they protect from industrial exploitation.",
 [("Homebound (Kalimdor only)", "Bound to Tauren sacred lands and natural sites"),
  ("Drifter (no hearth/bank)", "Wandering seeker of natural knowledge, no fixed home"),
  ("Intellect stacking gear", "Balance druids need massive spell power for moon/sun magic"),
  ("Pro-nature quests", "Fight the Venture Company across Barrens and Stonetalon")],
 "https://warcraft.wiki.gg/wiki/Druid"),

("Buccaneer", "Hunter", "Melee weaving survival", "Any", "Any",
 "Buccaneers are seafaring hunters who sail between ports, taming jungle beasts and wielding guns alongside rapiers. They live by the pirate's code, answering to no crown, with a parrot on one shoulder and a cutlass in hand.",
 [("Gun", "Pirates wield firearms for ranged combat on deck and shore"),
  ("Rapier/cutlass/harpoon", "Iconic pirate melee weapons"),
  ("Captain's hat ", "Every pirate captain earns their hat"),
  ("Renegade (can't equip quest rewards)", "Pirates take what they want, not quest handouts"),
  ("Parrot companion", "Tropical animal companions of the seafarer"),
  ("Jungle cat pet", "Tropical animal companions of the seafarer"),
  ("Rum drinker", "Pirates are famous for their love of grog"),
  ("Melee weaving talents", "Pirates are famous for their love of grog")],
 "https://warcraft.wiki.gg/wiki/Pirate"),

("Beastmaster", "Hunter", "Deep beast mastery", "Orc", "Any",
 "Beastmasters are orc hunters who devote their lives to commanding powerful wild beasts. They reject civilized firearms. Their pets are brothers, not tools — and their death is permanent.",
 [("No guns", "Beastmasters reject firearms, relying on bows and primal combat"),
  ("Beast fighting gear", "Wear trophies of the hunt as proof of mastery"),
  ("Wolf helm", "Iconic beastmaster headpiece showing pack leadership"),
  ("Mortal pets (pet death is permanent)", "Pet death is permanent — the bond is sacred and irreplaceable"),
  ("Leatherworking", "Craft armor from hides of their hunts"),
  ("Big game quests", "Hunt legendary apex predators across Azeroth"),
  ("Rare pets", "Seek out rare-spawn tameable beasts as trophies")],
 "https://warcraft.wiki.gg/wiki/Beastmaster_(Warcraft_III)"),

("Mountaineer", "Hunter", "Deep marksmanship", "Dwarf", "Any",
 "Mountaineers are dwarven rangers using custom-built rifles and massive axes. They seek out quests that bring glory to Ironforge and only use rewards from quests.",
 [("Self-made guns", "Dwarven weapons of choice — rifles and battleaxes"),
  ("Engineering", "Mountaineers engineer and maintain their own custom firearms"),
  ("2h axe", "All ranged weapons must be personally crafted"),
  ("Partisan (can't equip looted gear)", "Military-issued gear only — no battlefield scavenging"),
  ("Bear pet", "Mountain bears are traditional dwarven animal companions"),
  ("Ironforge loyalist quests", "Serve the dwarven kingdom of Ironforge")],
 "https://warcraft.wiki.gg/wiki/Mountaineer"),

("Spirit Champion", "Shaman", "Tank enhancement", "Orc", "Any",
 "Spirit Champions are orc shamans who channel ancestral power through melee combat. They are picky about the gear they wear.",
 [("Shield", "Defensive shields channeling ancestral totem power"),
  ("1200 armor at 30 / 3000 at 50", "Must reach armor thresholds proving defensive mastery"),
  ("No uncommon/green gear", "Spirit Champions wear unique shamanic ceremonial armor"),
  ("Rare Collector quests", "Gather rare artifacts honoring ancestral spirits"),
  ("Meditative Pauses", "Use /sit between fights to commune with spirits")],
 "https://warcraft.wiki.gg/wiki/Shaman"),

("Witch Doctor", "Shaman", "Totem restoration", "Troll", "Female",
 "Witch Doctors are female Troll shamans who practice voodoo magic, serving as healers and spiritual guides. They wear ritual masks and turtle shells, and create potent brews and elixirs.",
 [("Shell shield", "Ritual turtle shells serve as magical voodoo focuses"),
  ("Voodoo mask", "Masks channel ancestral spirits and voodoo magic"),
  ("Cursed amulet", "Dark jewelry binding hex magic"),
  ("Alchemy", "Witch Doctors brew voodoo potions, hexes, and healing tonics"),
  ("Cloth/Leather only", "Shamanic tradition favors lighter ritual garments"),
  ("Renegade (can't equip quest rewards)", "Witch Doctors operate outside formal tribal structures")],
 "https://warcraft.wiki.gg/wiki/Witch_doctor"),

("Spiritwalker", "Shaman", "Deep elemental", "Tauren", "Any",
 "Spiritwalkers are Tauren shamans who travel solo & far to find kindred spirits. They carry lanterns to light shadowed spirit paths and craft all their own gear.",
 [("1h axe", "Shamanic ritual axes for both combat and ceremony"),
  ("Lantern off-hand", "Magical lanterns illuminate paths between the spirit and physical worlds"),
  ("Leatherworking", "Spiritwalkers craft their own armor from natural hides"),
  ("Self-made gear", "All equipment must be personally crafted through tradition"),
  ("Wander-the-land quests", "They travel seeking spiritual knowledge and elemental sites")],
 "https://warcraft.wiki.gg/wiki/Spiritwalker"),

("Exemplar", "Paladin", "Deep holy", "Human", "Female",
 "Exemplars are human paladins who embody the ideals of the Church of the Holy Light. They are living symbols of faith and justice, wearing guild tabards and Stormwind colors as ordained protectors of the innocent.",
 [("Shield", "Holy warriors protect the innocent behind blessed shields"),
  ("Guild tabard", "Church tabards displaying their holy order affiliation"),
  ("Insignia of the Alliance", "Badge of rank within the Church hierarchy"),
  ("Partisan (can't equip looted gear)", "Holy warriors use sanctioned equipment only"),
  ("Mail/Plate only", "Divine armor channeling holy protection"),
  ("Stormwind hearthstone", "Divine armor channeling holy protection"),
  ("Stormwind loyalist quests", "Serve the kingdom of Stormwind exclusively")],
 "https://warcraft.wiki.gg/wiki/Paladin"),

("Templar", "Paladin", "Deep prot tank", "Human", "Male",
 "Templars are male human paladins of the Knights of the Silver Hand, sworn to protect the living against undead corruption. They carry the Argent Dawn's trinket as proof of their sacred anti-Scourge oath.",
 [("Sword/mace + shield", "Traditional Silver Hand knight weapons"),
  ("Argent Dawn trinket", "Carry the symbol of the anti-Scourge order at all times"),
  ("Homebound (Eastern Kingdoms)", "Oath-sworn to defend their homeland of Lordaeron/Eastern Kingdoms"),
  ("Anti-undead quests", "Their sacred purpose: fighting the Scourge in the Plaguelands")],
 "https://warcraft.wiki.gg/wiki/Knights_of_the_Silver_Hand"),

("Sister of Steel", "Paladin", "Deep retribution", "Dwarf", "Female",
 "Sisters of Steel are female dwarven blacksmiths devoted to forging weapons and armor. They are warrior-smiths who supply the army as well as themselves.",
 [("Blacksmithing", "Core identity — they must forge their own holy weapons and armor"),
  ("Self-made gear", "All equipment is personally smithed with divine light"),
  ("Ram mount", "Dwarven war rams are traditional paladin mounts"),
  ("Epic Hammer Journey quests", "Quest chain following the path to master blacksmithing")],
 "https://warcraft.wiki.gg/wiki/Paladin"),

("Priestess of the Moon", "Priest", "Spirit-based holy dps", "Night Elf", "Female",
 "Priestesses of the Moon are ancient spiritual leaders who ride sacred frostsaber mounts and wield Elune's divine power.",
 [("Robe", "Mystical moonsilk robes channeling Elune's light"),
  ("Spirit stacking (180/250)", "Moon magic is amplified through deep spirit attunement"),
  ("Partisan (can't equip looted gear)", "Priestesses use only sanctified equipment"),
  ("Frostsaber mount", "Sacred night elf mount connected to Elune's grace"),
  ("Spirit Tap + Starshards gameplay", "Boost damage with Spirit Tap, cast Starshards before buff fades")],
 "https://warcraft.wiki.gg/wiki/Priestess_of_the_Moon"),

("Apothecary", "Priest", "Deep discipline", "Undead", "Any",
 "Apothecaries are undead members of the Royal Apothecary Society. They are support units for the Forsaken army. They craft elixirs in the bowels of the Undercity.",
 [("Alchemy", "Core identity — they craft all elixirs, potions, and plagues"),
  ("Dagger + Robe", "Apothecary uniform: robes and small alchemical blades"),
  ("Herb pouch", "Carry harvested reagents for on-the-go potion brewing"),
  ("Homebound (Eastern Kingdoms)", "Bound to the Undercity and their apothecary laboratories"),
  ("Plague-brewer quest chain", "Follow the iconic 'New Plague' questline creating elixirs of suffering, pain, and agony"),
  ("Cockroach companion", "Undercity vermin serve as lab familiars")],
 "https://warcraft.wiki.gg/wiki/Royal_Apothecary_Society"),

("Shadow Hunter", "Priest", "Melee weaving shadow", "Troll", "Any",
 "Shadow Hunters are Troll spiritual leaders wielding voodoo curses and wearing ritual masks. They aren't traditional spellcasters - they can dish out powerful melee strikes. They serve as commanders rallying the Darkspear tribe.",
 [("No robes / wands", "Shadow Hunters wear tribal attire, not priestly vestments"),
  ("Staff then Pole at 44", "Heavy shamanic staves of power, transitioning to glaives"),
  ("Voodoo mask", "Ritual masks channel shadow and voodoo magic"),
  ("120 AP at 50", "Shadow Hunters are melee-capable, requiring attack power"),
  ("Darkspear Loyalist quests", "Serve the Darkspear tribe and gather ritual weapons"),
  ("Melee weaving gameplay", "Weave melee auto-attacks while channeling Mind Flay")],
 "https://warcraft.wiki.gg/wiki/Shadow_hunter"),

("Bloodmage", "Mage", "Deep fire", "Undead", "Female",
 "Bloodmages are mages  who turned to forbidden fire magic. They are master enchanters and travel with phoenix companions. They rely on their own enhancting and wear only white/grey gear.",
 [("Enchanting", "Bloodmages inscribe their own weapons with blood-fire runes"),
  ("Shadow or fire wand", "Wands channel hybrid shadow-fire energy"),
  ("Unholy weapon enchant", "Weapons imbued with unholy death magic"),
  ("White Knight (white/grey gear only)", "Ironic purity — plain gear serves as canvas for blood enchantments"),
  ("Drifter (no hearth/bank)", "Bloodmages belong to no kingdom, wandering between worlds"),
  ("Phoenix companion", "Al'ar-inspired phoenix familiars serve as magical companions")],
 "https://warcraft.wiki.gg/wiki/Blood_mage"),

("Mechano-Mage", "Mage", "Deep arcane", "Gnome", "Any",
 "Mechano-Mages are gnome arcane scholars who fuse engineering with magic, wearing progressively advanced goggles that enhance their arcane sight.",
 [("Engineering", "Core identity — they must engineer all their techno-arcane gear"),
  ("Goggles progression", "Enchanted goggles upgrade as arcane sight grows"),
  ("Renegade (can't equip quest rewards)", "Operating outside conventional gnome society as inventors"),
  ("Gadgetist quests", "Pursue mechanical and arcane innovations across Azeroth"),
  ("Mechanical companion", "Golem or mechanical squirrel familiars"),
  ("Pyroblast + Arcane Missiles gameplay", "Open fights with maximum burst, then regen mana")],
 "https://warcraft.wiki.gg/wiki/Gnome"),

("Warmage", "Mage", "Aoe-farmer", "Human", "Any",
 "Warmages are human battle-mages who serve in military campaigns, wielding sword in one hand and a staff in the other. They fight on the front lines alongside footmen using frost magic to crowd control.",
 [("Sword + Staff-like off-hand", "Military mages carry a blade and a focus into battle"),
  ("Armored ring", "Reinforced rings for close-quarters combat protection"),
  ("Footman (can't equip rare/epic items)", "Rank-and-file military gear — no special treatment"),
  ("Snow rabbit companion", "Frost affinity extends to arctic animal companions"),
  ("AoE farmer", "Pull big packs and burn them down with Blizzard and Frost Nova")],
 "https://warcraft.wiki.gg/wiki/Warmage"),
]

for ri, (name, base, spec, race, gender, lore, reqs, url) in enumerate(DATA, 2):
    fill = FILLS[base]
    reqs_col = "\n".join(f"- {r}" for r, _ in reqs)
    just_col = "\n".join(f"- {r}: {j}" for r, j in reqs)

    vals = [name, base, spec, race, gender, lore, reqs_col, just_col, url]
    for ci, val in enumerate(vals, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.font = BOLD if ci == 1 else BODY
        c.fill = fill
        c.alignment = WRAP
        c.border = BDR

    link_cell = ws.cell(row=ri, column=9)
    if link_cell.value:
        link_cell.hyperlink = link_cell.value
        link_cell.font = LINK

    ws.row_dimensions[ri].height = 105

import os
out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "HCE_Lore_Guide.xlsx")
wb.save(out)
print(f"Saved: {out}")
